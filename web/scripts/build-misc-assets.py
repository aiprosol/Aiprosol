#!/usr/bin/env python3
"""
Build supporting assets referenced by various deliverables:
  - tco-calculator.xlsx (for AI Tools Stack Starter Kit)
  - tool-cost-comparison.xlsx (for Global Business Automation Starter Pack)
  - tools-catalogue.csv + .json (for The AI Tools Vault)
"""

from pathlib import Path
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "private-products"

VIOLET = "8B5CF6"
LIGHT = "F4F1FA"
THIN = Border(
    left=Side(style="thin", color="D4D4D4"),
    right=Side(style="thin", color="D4D4D4"),
    top=Side(style="thin", color="D4D4D4"),
    bottom=Side(style="thin", color="D4D4D4"),
)


# ─────────────────────────────────────────────────────────────────────────
# 1. TCO Calculator — for the AI Tools Stack Starter Kit
# ─────────────────────────────────────────────────────────────────────────

def build_tco_calculator():
    wb = Workbook()
    ws = wb.active
    ws.title = "TCO Calculator"

    ws["A1"] = "Aiprosol · AI Stack TCO Calculator"
    ws["A1"].font = Font(size=18, bold=True, color=VIOLET)
    ws.merge_cells("A1:F1")

    ws["A2"] = "Plug in your usage. Outputs annual cost + cost-per-user."
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:F2")

    ws.column_dimensions["A"].width = 28
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16

    # Org assumptions
    ws["A4"] = "ORG ASSUMPTIONS"
    ws["A4"].font = Font(bold=True, color=VIOLET)
    for r, (label, default) in enumerate([
        ("Employees (FTE)", 25),
        ("Avg fully-loaded $/hour", 100),
        ("Active automation workflows", 12),
        ("Avg runs/day per workflow", 50),
    ], start=5):
        ws[f"A{r}"] = label
        ws[f"B{r}"] = default
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"].fill = PatternFill("solid", fgColor=LIGHT)

    # Tool grid
    ws["A10"] = "TOOL"
    ws["B10"] = "TIER / PLAN"
    ws["C10"] = "MONTHLY $"
    ws["D10"] = "ANNUAL $"
    ws["E10"] = "PER-USER $"
    ws["F10"] = "NOTES"
    for col in "ABCDEF":
        c = ws[f"{col}10"]
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=VIOLET)
        c.border = THIN

    tools = [
        ("Claude (Anthropic API)", "Pay-as-you-go", 30, "10M tokens/mo at $3/1M"),
        ("GPT-4o-mini (OpenAI)", "Pay-as-you-go", 8, "50M tokens/mo at $0.15/1M"),
        ("n8n self-hosted", "VPS $5/mo", 5, "Unmetered. Or n8n Cloud Starter $20."),
        ("HubSpot Free", "$0", 0, "Free up to 1M contacts. Upgrade only at scale."),
        ("Supabase Pro", "Pro plan", 25, "DB + auth + storage + pgvector"),
        ("Resend", "$20 + usage", 35, "100k emails for $35"),
        ("Cal.com self-hosted", "$0", 0, "Or Calendly Standard $10/mo"),
        ("PostHog Cloud", "Free tier", 0, "Free up to 1M events/mo"),
        ("Plain", "Plain Basic", 79, "Per seat $79/mo"),
        ("Notion", "$10 × 25 seats", 250, "Wiki + databases"),
        ("Google Workspace", "$14 × 25 seats", 350, "Business Standard"),
        ("Slack Pro", "$8 × 25 seats", 200, "Pro plan"),
        ("Stripe (billing)", "2.9% + 30c", 0, "Passthrough — varies by rev"),
        ("Mercury", "$0", 0, "US banking, free"),
    ]
    for i, (name, plan, monthly, note) in enumerate(tools, start=11):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = plan
        ws[f"C{i}"] = monthly
        ws[f"D{i}"] = f"=C{i}*12"
        ws[f"E{i}"] = f"=D{i}/B5"
        ws[f"F{i}"] = note
        for col in "ABCDEF":
            ws[f"{col}{i}"].border = THIN
            if (i - 10) % 2 == 0:
                ws[f"{col}{i}"].fill = PatternFill("solid", fgColor=LIGHT)
        ws[f"C{i}"].number_format = "$#,##0"
        ws[f"D{i}"].number_format = "$#,##0"
        ws[f"E{i}"].number_format = "$#,##0.00"

    total_row = 11 + len(tools)
    ws[f"A{total_row}"] = "TOTAL FIXED"
    ws[f"A{total_row}"].font = Font(bold=True, size=12, color=VIOLET)
    ws[f"C{total_row}"] = f"=SUM(C11:C{total_row-1})"
    ws[f"D{total_row}"] = f"=SUM(D11:D{total_row-1})"
    ws[f"E{total_row}"] = f"=D{total_row}/B5"
    for col in "CDE":
        ws[f"{col}{total_row}"].font = Font(bold=True, size=12)
        ws[f"{col}{total_row}"].number_format = "$#,##0"

    # Break-even section
    be_row = total_row + 3
    ws[f"A{be_row}"] = "BREAK-EVEN GUIDANCE"
    ws[f"A{be_row}"].font = Font(bold=True, color=VIOLET, size=12)
    rules = [
        "If workflows × runs/day > 5000: migrate from Zapier to Make",
        "If workflows × runs/day > 50000: migrate from Make to n8n self-hosted",
        "If LLM tokens/mo > 100M: consider Together.ai / Groq for cost arbitrage",
        "If seat-cost > 25% of total: re-evaluate per-seat tools (Notion, Slack alternatives)",
    ]
    for i, r in enumerate(rules, start=be_row + 1):
        ws[f"A{i}"] = r

    ws["A1"].alignment = Alignment(vertical="center")
    out = OUT / "tco-calculator.xlsx"
    wb.save(out)
    return out


# ─────────────────────────────────────────────────────────────────────────
# 2. Tool Cost Comparison — Zapier vs Make vs n8n
# ─────────────────────────────────────────────────────────────────────────

def build_tool_cost_comparison():
    wb = Workbook()
    ws = wb.active
    ws.title = "Workflow Tool Comparison"

    ws["A1"] = "Zapier vs Make vs n8n · Cost Comparison"
    ws["A1"].font = Font(size=18, bold=True, color=VIOLET)
    ws.merge_cells("A1:F1")

    ws["A2"] = "Enter your volume below. Costs auto-calculate."
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:F2")

    ws.column_dimensions["A"].width = 28
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16

    ws["A4"] = "INPUTS"
    ws["A4"].font = Font(bold=True, color=VIOLET)
    ws["A5"] = "Workflows in production"
    ws["B5"] = 15
    ws["A6"] = "Average runs/day per workflow"
    ws["B6"] = 50
    ws["A7"] = "Total monthly runs"
    ws["B7"] = "=B5*B6*30"
    ws["B7"].number_format = "#,##0"

    ws["A9"] = "TOOL COST AT YOUR VOLUME"
    ws["A9"].font = Font(bold=True, color=VIOLET)

    headers = ["Tool", "Monthly cost", "Annual cost", "Cost / 1000 runs", "Notes"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=10, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=VIOLET)
        c.border = THIN

    # Pricing models (approximate Jan 2026):
    # Zapier Team: $69/mo + $0.0085/task above 2000
    # Make Core: $9/mo + $9 per 10k ops
    # n8n self-hosted: $5/mo VPS
    # n8n Cloud: $20/mo Starter, $50 Pro
    ws["A11"] = "Zapier Team"
    ws["B11"] = "=69+IF(B7>2000,(B7-2000)*0.0085,0)"
    ws["A12"] = "Make Core"
    ws["B12"] = "=9+ROUNDUP(B7/10000,0)*9"
    ws["A13"] = "n8n self-hosted"
    ws["B13"] = 5
    ws["A14"] = "n8n Cloud Pro"
    ws["B14"] = 50

    for r in range(11, 15):
        ws[f"C{r}"] = f"=B{r}*12"
        ws[f"D{r}"] = f"=B{r}/(B7/1000)"
        ws[f"B{r}"].number_format = "$#,##0"
        ws[f"C{r}"].number_format = "$#,##0"
        ws[f"D{r}"].number_format = "$#,##0.00"
        for col in "ABCDE":
            ws[f"{col}{r}"].border = THIN
            if r % 2 == 1:
                ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=LIGHT)

    ws["E11"] = "Most polished UX, most expensive per run"
    ws["E12"] = "Best value at mid-volume, visual builder"
    ws["E13"] = "Unmetered. Requires dev to set up"
    ws["E14"] = "Same product, hosted by n8n"

    # Recommendation
    ws["A17"] = "RECOMMENDATION"
    ws["A17"].font = Font(bold=True, color=VIOLET, size=12)
    ws["A18"] = '=IF(B7<3000,"Zapier — simplicity wins below 100 runs/day",IF(B7<30000,"Make — best $/run + features at this scale","n8n self-hosted — unmetered economics win at this scale"))'
    ws["A18"].font = Font(bold=True, size=12)
    ws.merge_cells("A18:E18")

    out = OUT / "tool-cost-comparison.xlsx"
    wb.save(out)
    return out


# ─────────────────────────────────────────────────────────────────────────
# 3. AI Tools Catalogue — CSV + JSON
# ─────────────────────────────────────────────────────────────────────────

# 100 curated entries spanning the 23 categories. Real tools we'd
# recommend — honest verdicts, not hype.

CATALOGUE = [
    # Conversational AI — frontier
    ("Claude", "Conversational AI", "Anthropic frontier LLM. Best for accuracy + long context.", "$3-15/1M tokens", "PICK", "Use for customer-facing + judgement-sensitive."),
    ("GPT-4o", "Conversational AI", "OpenAI's flagship multimodal model.", "$2.50-10/1M tokens", "PICK", "Best ecosystem + integrations."),
    ("GPT-4o-mini", "Conversational AI", "Cheap workhorse for bulk classification.", "$0.15-0.60/1M tokens", "PICK", "Use for high-volume backend grunt work."),
    ("Gemini 1.5 Pro", "Conversational AI", "Google's frontier model. Free tier via AI Studio.", "Free tier + $1.25/1M", "GOOD", "Best if you live in Google Workspace."),
    ("Claude Haiku", "Conversational AI", "Anthropic's small + fast model.", "$0.80/1M tokens", "PICK", "Cheaper Claude for routine tasks."),
    # Conversational AI — open
    ("Llama 3.3 70B", "Open-source LLM", "Meta's flagship open model.", "Free + hosting", "PICK", "Self-host or via Groq/Together for cheap inference."),
    ("Mistral Large 2", "Open-source LLM", "Strong European open model.", "Pay-per-use", "GOOD", "Better for European data residency."),
    ("DeepSeek V3", "Open-source LLM", "Strong reasoning + math, open weights.", "Free + hosting", "GOOD", "Surprisingly capable for math/code."),
    ("Qwen 2.5 72B", "Open-source LLM", "Alibaba's open model, multilingual.", "Free + hosting", "GOOD", "Best for non-English at scale."),
    # Code gen
    ("Claude Code", "Code generation", "Anthropic's CLI agent for code.", "Bundled with Claude API", "PICK", "Best agentic coding tool."),
    ("Cursor", "Code generation", "AI-first IDE.", "$20/mo", "PICK", "Best balance of speed + control."),
    ("Windsurf", "Code generation", "Codeium's agentic IDE.", "$15/mo", "GOOD", "Alternative to Cursor."),
    ("GitHub Copilot", "Code generation", "Embedded in VS Code / JetBrains.", "$10/mo", "GOOD", "Adequate. Falling behind frontier."),
    ("Aider", "Code generation", "Terminal-native pair programming.", "Free (BYO API)", "GOOD", "For terminal-first devs."),
    # Image gen
    ("Midjourney", "Image generation", "Quality leader for stylised images.", "$10-60/mo", "PICK", "Use for marketing creative."),
    ("DALL-E 3", "Image generation", "Best prompt-following.", "Pay-per-use", "GOOD", "Embedded in ChatGPT Plus."),
    ("Stable Diffusion 3.5", "Image generation", "Open weights, full control.", "Free + GPU", "GOOD", "When you need control or volume."),
    ("Ideogram", "Image generation", "Best for text-in-image.", "Free tier + $8/mo", "GOOD", "Posters, logos with text."),
    # Video gen
    ("Runway Gen-3", "Video generation", "Best for editing-style video work.", "$15-95/mo", "PICK", "Production-ready clips."),
    ("Veo (Google)", "Video generation", "Best realism.", "Limited access", "WATCH", "Wait for general release."),
    ("Pika", "Video generation", "Animation + stylised video.", "$10-95/mo", "GOOD", "Faster than Runway for sketches."),
    # Voice cloning
    ("ElevenLabs", "Voice cloning", "Quality + voice library.", "$5-330/mo", "PICK", "Best quality on the market."),
    ("Play.ht", "Voice cloning", "Most stable for long-form.", "$39-99/mo", "GOOD", "Reliable for podcasts."),
    ("Cartesia", "Voice cloning", "Latency leader (real-time).", "$10-99/mo", "PICK", "For live voice agents."),
    # Voice transcription
    ("Deepgram", "Voice transcription", "Best for live transcription.", "Pay-per-use", "PICK", "Real-time use cases."),
    ("Whisper", "Voice transcription", "OpenAI open-source transcription.", "Free + GPU", "PICK", "Batch transcription, self-host."),
    ("Otter", "Voice transcription", "Meeting-focused.", "$10-30/mo", "GOOD", "For meeting notes specifically."),
    # RAG / orchestration
    ("LangChain", "RAG orchestration", "Python framework for LLM apps.", "Free", "GOOD", "Flexible but heavy. Pick selectively."),
    ("LlamaIndex", "RAG orchestration", "RAG-focused framework.", "Free", "PICK", "Best RAG-specific tool."),
    ("Vellum", "LLM ops", "Production prompt management.", "$50-300/mo", "GOOD", "For teams with >5 prod prompts."),
    ("LangSmith", "LLM ops", "LangChain's observability platform.", "Free + paid", "GOOD", "If you use LangChain."),
    # Vector DBs
    ("pgvector (Supabase)", "Vector DB", "Postgres + vectors.", "$25/mo Pro", "PICK", "Best for SMB. One DB, fewer integrations."),
    ("Pinecone", "Vector DB", "Specialist managed vector DB.", "$70+/mo", "GOOD", "Only when 100M+ vectors."),
    ("Weaviate", "Vector DB", "Hybrid search-friendly.", "Free + paid", "GOOD", "Self-host for control."),
    # LLM hosting
    ("Groq", "LLM inference", "Fastest inference (LPU hardware).", "Pay-per-use", "PICK", "When latency matters most."),
    ("Together AI", "LLM inference", "Wide model selection.", "Pay-per-use", "PICK", "Cheap alternative to OpenAI."),
    ("Fireworks AI", "LLM inference", "Fast + serverless.", "Pay-per-use", "GOOD", "Good for fine-tuned models."),
    ("Replicate", "LLM inference", "Image + video model hosting.", "Pay-per-use", "GOOD", "For media model variety."),
    # Orchestration
    ("n8n", "Workflow automation", "Open-source visual workflow tool.", "$5/mo VPS or $20+ cloud", "PICK", "Best for engineers + ops teams."),
    ("Make.com", "Workflow automation", "Visual workflow at SMB scale.", "$9-29/mo", "PICK", "Best for non-developer ops teams."),
    ("Zapier", "Workflow automation", "Most integrations (5000+).", "$20-69+/mo", "GOOD", "Use for niche tool integrations."),
    ("Pipedream", "Workflow automation", "Code-first serverless workflows.", "Free + paid", "GOOD", "If you prefer code over visual."),
    # No-code app builders
    ("v0 (Vercel)", "No-code/landing", "AI landing pages + components.", "$20-50/mo", "PICK", "For landing pages + UI."),
    ("Lovable", "No-code/app", "AI full-stack app builder.", "$20-99/mo", "GOOD", "Whole-app prototypes."),
    ("Replit", "No-code/app", "AI-assisted full-stack.", "$15-25/mo", "GOOD", "Quick prototypes + agent workflows."),
    ("Bolt.new", "No-code/app", "AI app builder.", "$20-100/mo", "WATCH", "Newer entrant, evolving."),
    # Forms
    ("Tally", "Forms", "Free, polished forms.", "Free + $24/mo", "PICK", "Best free tier."),
    ("Typeform", "Forms", "Beautiful forms.", "$25-83/mo", "GOOD", "For brand-sensitive surveys."),
    ("Fillout", "Forms", "Logic-heavy forms.", "$20-50/mo", "GOOD", "When you need branching."),
    # Email / outreach
    ("Resend", "Email API", "Modern email API.", "$20+ + usage", "PICK", "API-first email sending."),
    ("Lemlist", "Cold outreach", "Cold email + personalisation.", "$59-99/mo", "GOOD", "If you do cold outreach at scale."),
    ("Customer.io", "Email automation", "Behavioural email + sequences.", "$100+/mo", "PICK", "For lifecycle email."),
    ("Loops", "Email automation", "Startup-friendly transactional + lifecycle.", "$49+/mo", "GOOD", "Easy onboarding."),
    # CRM
    ("HubSpot", "CRM", "Free CRM with strong marketing.", "Free + $20-3600/mo", "PICK", "Free tier is the default SMB choice."),
    ("Pipedrive", "CRM", "Sales-pipeline-first CRM.", "$19-99/mo", "PICK", "Best for sales-led orgs."),
    ("Attio", "CRM", "Modern CRM.", "$29-99/mo", "GOOD", "Like Pipedrive but built in 2024."),
    ("Folk", "CRM", "Network-style CRM.", "$20+/mo", "GOOD", "For consultancy / agency use."),
    # Support
    ("Plain", "Customer support", "Engineer-friendly B2B support.", "$79+/mo per seat", "PICK", "For SaaS / technical product."),
    ("Intercom", "Customer support", "Multi-channel + AI agent.", "$39-139+/mo", "GOOD", "For B2C SaaS / mass market."),
    ("Help Scout", "Customer support", "Inbox-style support.", "$25-65/mo", "GOOD", "Email-centric teams."),
    # Analytics
    ("PostHog", "Product analytics", "Analytics + recordings + flags.", "Free + paid", "PICK", "Best SMB choice. One tool, multiple jobs."),
    ("Amplitude", "Product analytics", "Cohort-focused analytics.", "Free + $50k+/yr", "GOOD", "Enterprise cohorts."),
    ("Mixpanel", "Product analytics", "Event-based analytics.", "Free + paid", "GOOD", "Alternative to Amplitude."),
    # Observability
    ("Sentry", "Error tracking", "Best error monitoring.", "Free + paid", "PICK", "Industry standard for error tracking."),
    ("Axiom", "Logs", "Modern log management.", "Free + usage", "PICK", "Cheaper than Datadog."),
    ("Highlight.io", "Session replay", "Full-session recording.", "Free + paid", "GOOD", "If you need session replays."),
    # Document processing
    ("Unstructured", "Document parsing", "Best PDF/document parser.", "Pay-per-use", "PICK", "For RAG document prep."),
    ("Mathpix", "Math OCR", "Math-aware OCR.", "Free + paid", "GOOD", "Niche but excellent at math."),
    ("Reducto", "Tables OCR", "Best PDF table extraction.", "Pay-per-use", "PICK", "Better than Claude on tables."),
    # Productivity
    ("Notion", "Wiki / docs", "Wiki + databases + project tracking.", "$10/seat/mo", "PICK", "SMB default."),
    ("Linear", "Issue tracking", "Engineering-first issue tracker.", "$8-14/seat/mo", "PICK", "Best for product/eng teams."),
    ("Granola", "Meeting notes", "AI meeting notes.", "$14-25/mo", "PICK", "Best meeting notes app."),
    ("Cap", "Loom replacement", "Open-source screen recording.", "Free + paid", "GOOD", "Privacy-respecting alt to Loom."),
    # Marketing
    ("Ahrefs", "SEO", "Best SEO research tool.", "$108-449/mo", "PICK", "Industry standard for SEO."),
    ("Frase", "Content / SEO", "Content briefs from SERP.", "$15-115/mo", "GOOD", "For content teams."),
    ("Beehiiv", "Newsletters", "Newsletter platform.", "Free + paid", "GOOD", "Best newsletter tool right now."),
    ("ConvertKit (Kit)", "Newsletters", "Creator-focused email.", "Free + $25+/mo", "GOOD", "Creator alternative."),
    # Finance
    ("Stripe", "Payments + billing", "Standard SMB billing platform.", "2.9% + 30c", "PICK", "Default for online billing."),
    ("Mercury", "Banking", "Startup-friendly US banking.", "Free", "PICK", "For US-incorporated startups."),
    ("Wise Business", "Banking", "International banking.", "Free", "PICK", "For multi-currency businesses."),
    ("Pulley", "Cap table", "Cap table mgmt.", "Free + paid", "GOOD", "For equity tracking."),
    # Security
    ("1Password", "Password mgr", "Team password manager.", "$8-12/seat/mo", "PICK", "Standard for SMB security."),
    ("Vanta", "SOC2 / compliance", "SOC 2 automation.", "$10k+/yr", "PICK", "For SOC 2 audits."),
    ("Drata", "SOC2 / compliance", "Alt to Vanta.", "$10k+/yr", "GOOD", "Competitive with Vanta."),
    # Calendar
    ("Cal.com", "Booking", "Open-source Calendly.", "Free self-host + paid", "PICK", "For technical teams."),
    ("Calendly", "Booking", "Industry standard.", "Free + $10-16/mo", "PICK", "For non-technical teams."),
    # AI agents / specialized
    ("LangGraph", "Agent framework", "LangChain's agent orchestration.", "Free", "GOOD", "If using LangChain."),
    ("CrewAI", "Agent framework", "Multi-agent orchestration.", "Free", "WATCH", "Earlier-stage, evolving."),
    ("E2B", "Agent sandbox", "Code interpreter sandboxes.", "Pay-per-use", "PICK", "For code-executing agents."),
    ("Browserbase", "Browser agent", "Headless browser for AI.", "Pay-per-use", "PICK", "For browser-using agents."),
    ("Modal", "Serverless GPU", "Serverless GPU for AI.", "Pay-per-use", "PICK", "For custom AI workloads."),
    # Niche but high-value
    ("Apify", "Web scraping", "Web scraping platform.", "Free + paid", "GOOD", "For scraping at scale."),
    ("Firecrawl", "Web crawling", "AI-friendly crawler.", "$19+/mo", "PICK", "For RAG document collection."),
    ("Datalab", "OCR / parsing", "Marker open-source PDF.", "Pay-per-use", "GOOD", "Self-host alternative to Unstructured."),
    # AVOID listings (transparent honesty)
    ("Generic AI sales agent X", "Sales automation", "Often hallucinates outreach details.", "$199+/mo", "AVOID", "Use as agent-assisted, not agent-driven."),
    ("Heavy enterprise BI tool", "Analytics", "Costly + slow for SMB scale.", "$50k+/yr", "AVOID", "PostHog covers 90% of SMB needs."),
    ("Single-vendor AI suite", "Various", "Lock-in risk + variable quality.", "$200+/mo", "AVOID", "Build with best-of-breed instead."),
    ("AI customer-support agent", "Support", "Hallucination + brand-tone drift.", "$300+/mo", "AVOID", "Use AI for triage; humans close."),
    # Hidden gems
    ("Tursо (libSQL)", "Database", "Edge SQLite.", "Free + paid", "GEM", "Replicated SQLite at the edge."),
    ("Trigger.dev", "Background jobs", "Modern background jobs.", "Free + paid", "GEM", "Better DX than serverless functions."),
    ("Inngest", "Event-driven backend", "Event-first backend.", "Free + paid", "GEM", "For event-driven workflows."),
    ("Defer", "Background jobs", "Background jobs + workflows.", "Free + paid", "GEM", "Worth watching."),
    ("Webstudio", "Visual web builder", "Open-source visual web.", "Free + paid", "GEM", "Webflow alternative."),
    ("Raycast", "Mac productivity", "Spotlight + AI + scripts.", "Free + $8/mo Pro", "GEM", "Most underrated Mac tool."),
    ("Arc / Dia", "Browser", "Browser built around AI.", "Free", "GEM", "Worth trying as primary."),
]


def build_catalogue_files():
    csv_path = OUT / "ai-tools-catalogue.csv"
    json_path = OUT / "ai-tools-catalogue.json"

    # CSV
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Tool", "Category", "Summary", "Pricing", "Verdict", "Notes"])
        for row in CATALOGUE:
            w.writerow(row)

    # JSON
    json_data = [
        {
            "tool": r[0],
            "category": r[1],
            "summary": r[2],
            "pricing": r[3],
            "verdict": r[4],
            "notes": r[5],
        }
        for r in CATALOGUE
    ]
    with json_path.open("w", encoding="utf-8") as f:
        json.dump({
            "version": "1.0",
            "updated": "2026-05-15",
            "total": len(CATALOGUE),
            "tools": json_data,
        }, f, indent=2)

    return csv_path, json_path


if __name__ == "__main__":
    print("Building supporting assets...")
    f1 = build_tco_calculator()
    print(f"  ✓ {f1.name} ({f1.stat().st_size // 1024} KB)")
    f2 = build_tool_cost_comparison()
    print(f"  ✓ {f2.name} ({f2.stat().st_size // 1024} KB)")
    f3, f4 = build_catalogue_files()
    print(f"  ✓ {f3.name} ({f3.stat().st_size // 1024} KB)")
    print(f"  ✓ {f4.name} ({f4.stat().st_size // 1024} KB)")
    print(f"  catalogue: {len(CATALOGUE)} tools")
    print("Done.")
